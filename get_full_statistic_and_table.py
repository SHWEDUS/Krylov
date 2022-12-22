import csv
import math

import matplotlib
import matplotlib.pyplot as plt
import numpy as np

import openpyxl
from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter

from jinja2 import Environment, FileSystemLoader
import pdfkit
import doctest

SMALL_SIZE = 6
SIZE = 8
matplotlib.use('TkAgg')
plt.rcParams.update({'font.size': 8})

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}


class DataSet:
    """Класс для создания датасета с полученными вакансиями из csv файла

    Attributes:
        vacancies (list[list[str]]): Получаем вакансии в нужном нам виде, нужных столбцов
    """

    def __init__(self, file_name):
        """Инициализирует объект DataSet, для работы получает название csv файла, с которым мы работаем

        Args:
             file_name (str): Название csv файла
        """
        self.file_name = file_name
        self.vacancies = DataSet.parser_csv(file_name)

    @staticmethod
    def parser_csv(file_name):
        """Функция прасера csv, из неё мы получаем все наши вакансии в нужном виде

        Attributes:
            file_name(str): Название csv файла
        Returns:
            vacancies(list[list[str]]): Вакансии в нужном виде нужных столбцов
        """
        title, data = DataSet.csv_reader(file_name)
        result = DataSet.csv_filer(title, data)
        vacancies = DataSet.get_vacancies(result, title)
        return vacancies

    @staticmethod
    def get_vacancies(result, title):
        """Фукция получения вакансий нужных столбцов
        Attributes:
            result(list[list[str]]): Исходный DataSet со всеми вакансиями
            title(list[str]): Названия столбцов в csv файле
        Returns:
            vacancies(list[list[str]]: Вакансии в нужном виде нужных столбцов
        """
        vacancies = []
        if result:
            for row in result:
                temp = {}
                for i in range(len(row)):
                    temp[title[i]] = row[i]
                vacancies.append(
                    Vacancies(temp['name'], Salary(temp['salary_from'], temp['salary_to'], temp['salary_currency']),
                              temp['area_name'], temp['published_at']))
            return vacancies
        else:
            return vacancies

    @staticmethod
    def csv_reader(file_name):
        """Функция считывания данных из csv файла, получение названия столбцов и самих данных
        Attributes:
            file_name(str): Название csv файла
        Returns:
             head_arr(list[str]): Названия столбцов csv Файла
             data_arr(list[list[str]]): Исходные данные из csv файла (вакансии)
        """
        with open(file_name, encoding="utf-8-sig") as file:
            reader = csv.reader(file)
            data_arr = list(reader)
            if not data_arr:
                head_arr = data_arr
                return head_arr, data_arr
            else:
                head_arr = data_arr[0]
                data_arr.pop(0)
                return head_arr, data_arr

    @staticmethod
    def csv_filer(list_naming, reader):
        """Функция считывающая данные из csv, а также удаляющая пустые и содержащие None строки
        Attributes:
            list_naming(list[str]): Названия столбцов
            reader(list[list[str]]):  Исходные данные из csv файла (вакансии)
        Returns:
             result(list[list[str]]): Исходный DataSet со всеми вакансиями
        """
        head_array, data_array = list_naming, reader
        if data_array == [] and head_array == []:
            data_array = []
            return data_array
        for i in range(len(data_array)):
            if '' in data_array[i] or len(data_array[i]) < len(head_array):
                data_array[i].clear()
        data_array = list(filter(None, data_array))

        return data_array


class Salary:
    """Класс инициализирующий данных о зарплатах
    Attributes:
        salary_from(int): Нижняя граница оклада вакансии
        salary_to(int): Верхняя граница оклада вакансии
        salary_currency(str): Валюта оклада вакансии
    """
    def __init__(self, salary_from, salary_to, salary_currency):
        """Получаем данных о зарплатах для вакансии
        Args:
            salary_from(int or str or float): Нижняя граница оклада вакансии
            salary_to(int or str or float): Верхняя граница оклада вакансии
            salary_currency(str): Валюта оклада вакансии
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency


class Vacancies:
    """Класс инициализирующий данные о вакансиях
    Attributes:
        name(str): Название вакансии
        salary: Данные о зарпалатах по вакансии
        area_name(str): Место работы
        published_at(str): Дата публикации вакансии
    """
    def __init__(self, name, salary, area_name, published_at):
        """Получаем нужные нам данные о вакансии
        Args:
            name(str): Название вакансии
            salary(Salary(list[int or string or float]): Данные о зарпалатах по вакансии
            area_name(str): Место работы
            published_at(str): Дата публикации вакансии
        """
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at


class Report:
    """Класс создания репорта статистики и таблицы(в зависимости от выибора метода печати) в формате PDF
    Attributes:
        salary_year(dict{int: int}): Средняя заплата всех вакансий по годам
        count_year(dict{int: int}): Количество вакансий по годам
        salary_year_prof(dict{int: int}): Зарпалата выбранной профессии по годам
        count_year_prof(dict{int: int}): Колличество вакансий выбранной пофесиий по годам
        salary_dict_by_city(dict{str: int}): Средняя зарплата по городам,
        count_vacancies_dict_by_city(dict{str: int or str}): Количество вакансий по городам
        new_count(dict{str: int or str}): Дополнитенльный словарь для диаграммы по городам
        prof_name(str): Название профессии
        choosing_type(str): Выбор пользователя по поводу печати
    """
    def __init__(self, salary_year, count_year, salary_year_prof, count_year_prof, salary_dict_by_city,
                 count_vacancies_dict_by_city, new_count, prof_name, choosing_type):
        """Инициализация данных для статистики вакансий и выбранной профессии
        Args:
            salary_year(dict{int: int}): Средняя заплата всех вакансий по годам
            count_year(dict{int: int}): Количество вакансий по годам
            salary_year_prof(dict{int: int}): Зарпалата выбранной профессии по годам
            count_year_prof(dict{int: int}): Колличество вакансий выбранной пофесиий по годам
            salary_dict_by_city(dict{str: int}): Средняя зарплата по городам,
            count_vacancies_dict_by_city(dict{str: int}): Количество вакансий по городам
            new_count(dict{str: int}): Дополнитенльный словарь для диаграммы по городам
            prof_name(str): Название профессии
            choosing_type(str): Выбор пользователя по поводу печати
        """
        self.salary_year = salary_year
        self.count_year = count_year
        self.salary_year_prof = salary_year_prof
        self.count_year_prof = count_year_prof
        self.salary_dict_by_city = salary_dict_by_city
        self.count_vacancies_dict_by_city = count_vacancies_dict_by_city
        self.generate_excel = Report.generate_excel(salary_year, count_year, salary_year_prof, count_year_prof,
                                                    salary_dict_by_city, count_vacancies_dict_by_city, prof_name)

        self.sheet_1_headers = ['Год', 'Средняя зарпоата', 'Средняя зарплата - ' + prof_name, 'Количество вакансий',
                                'Количество вакансий - ' + prof_name]
        self.sheet_2_headers = ['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']
        sheet_1_columns = [list(salary_year.keys()), list(salary_year.values()), list(salary_year_prof.values()),
                           list(count_year.values()), list(count_year_prof.values())]
        self.sheet_1_rows = self.get_rows(sheet_1_columns)
        sheet_2_columns = [list(salary_dict_by_city.keys()), list(salary_dict_by_city.values()),
                           ['' for _ in salary_dict_by_city.keys()], list(salary_dict_by_city.keys()),
                           list(count_vacancies_dict_by_city.values())]
        self.sheet_2_rows = self.get_rows(sheet_2_columns)
        self.generate_pdf = Report.generate_pdf(prof_name, self.sheet_1_headers, self.sheet_2_headers,
                                                self.sheet_1_rows,
                                                self.sheet_2_rows, choosing_type)
        self.generate_image = Report.generate_image(salary_year, count_year, salary_year_prof, count_year_prof,
                                                    salary_dict_by_city, new_count, prof_name)

    def get_rows(self, columns: list):
        """Функция для создания строк таблицы вывода статистики
        Attributes:
            columns(list[list[int or string or float]]): Колонки заданные в таблице статистики
        Returns:
            rows(list[list[str]]): Полученные строки для таблицы статистики
        """
        rows = [['' for _ in range(len(columns))] for _ in range(len(columns[0]))]
        for col in range(len(columns)):
            for cell in range(len(columns[col])):
                rows[cell][col] = columns[col][cell]
        return rows

    @staticmethod
    def generate_excel(salary_year, count_year, salary_year_prof, count_year_prof, salary_dict_by_city,
                       count_vacancies_dict_by_city, profession):
        """Функция генератор таблицы в формате xlxs
        Attributes:
            salary_year(dict{int: int}): Средняя заплата всех вакансий по годам
            count_year(dict{int: int}): Количество вакансий по годам
            salary_year_prof(dict{int: int}): Зарпалата выбранной профессии по годам
            count_year_prof(dict{int: int}): Колличество вакансий выбранной пофесиий по годам
            salary_dict_by_city(dict{str: int}): Средняя зарплата по городам,
            count_vacancies_dict_by_city(dict{str: int or str}): Количество вакансий по городам
            profession(str): Название професии
        Returns:
             book(Workbook): Созданная таблица, чтобы потом сохранить её в нужном нам имени
        """
        book = openpyxl.Workbook()
        sheet_year = book.active
        sheet_year.title = 'Статистика по годам'
        sheet_year['A1'] = 'Год'
        sheet_year['B1'] = 'Средняя зарплата'
        sheet_year['C1'] = f'Средняя зарплата - {profession}'
        sheet_year['D1'] = 'Количество вакансий'
        sheet_year['E1'] = f'Количество вакансий - {profession}'

        column_letter_year = 'ABCDE'
        for letter in column_letter_year:
            sheet_year[f'{letter}1'].font = Font(bold=True)

        sheet_city = book.create_sheet('Статистика по городам')
        sheet_city['A1'] = 'Город'
        sheet_city['B1'] = 'Уровень зарплат'
        sheet_city['D1'] = 'Город'
        sheet_city['E1'] = 'Доля вакансий'

        column_letter_city = 'ABDE'
        for letter in column_letter_city:
            sheet_city[f'{letter}1'].font = Font(bold=True)

        for row, (key, value) in enumerate(salary_year.items(), start=2):
            sheet_year[f'A{row}'] = key
        year = [salary_year, salary_year_prof, count_year, count_year_prof]
        Report.fill_column(sheet_year, year, column_letter_year[1:])

        for row, (key, value) in enumerate(salary_dict_by_city.items(), start=2):
            sheet_city[f'A{row}'] = key
            sheet_city[f'D{row}'] = key
            sheet_city[f'B{row}'] = value

        for row, (key, value) in enumerate(count_vacancies_dict_by_city.items(), start=2):
            sheet_city[f'E{row}'] = value

        thin = Side(border_style='thin', color='000000')

        Report.get_width_cells(sheet_year, 'C')
        Report.get_width_cells(sheet_city, '')

        Report.get_border(sheet_year, thin)
        Report.get_border(sheet_city, thin)

        return book

    @staticmethod
    def fill_column(sheet, data, letters):
        """Функция заполнения коллонок таблицы статистики данными
        Attributes:
            sheet(Workbook): Созданный лист таблицы
            data(list[{items}]): Данные, которые нужно записать в таблицу
            letters(str): Нужные коллонки таблицы в буквенном формате
        """
        for i, item in enumerate(data):
            for row, (key, value) in enumerate(item.items(), start=2):
                sheet[f'{letters[i]}{row}'] = value

    @staticmethod
    def get_width_cells(sheet, letter):
        """Функция автоматической регулировки ширины коллонки
        Attributes:
            sheet(Workbook): Созданный лист таблицы
            letter(str): Буква колонки-пробела
        """
        column_widths = []
        for row in sheet:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell.value)) > column_widths[i]:
                        column_widths[i] = len(str(cell.value))
                else:
                    column_widths += [len(str(cell.value))]

        for i, column_width in enumerate(column_widths, 1):
            if letter == '':
                sheet.column_dimensions[get_column_letter(i)].width = (column_width + 1)
            else:
                if get_column_letter(i) == letter:
                    sheet.column_dimensions[get_column_letter(i)].width = 1.5
                else:
                    sheet.column_dimensions[get_column_letter(i)].width = (column_width + 2)

    @staticmethod
    def get_border(sheet, font):
        """Функция создания линий гранц таблицы
        Attributes:
            sheet(Workbook): Созданный лист таблицы
            font(Side): Толщина линии
        """
        for row in sheet.columns:
            for cell in row:
                if cell.value:
                    cell.border = Border(top=font, left=font, right=font, bottom=font)

    @staticmethod
    def generate_image(salary_year, count_year, salary_year_prof, count_year_prof,
                       salary_dict_by_city, count_vacancies_dict_by_city, prof_name):
        """Создание файла картинки диаграмм
        Attributes:
            salary_year(dict{int: int}): Средняя заплата всех вакансий по годам
            count_year(dict{int: int}): Количество вакансий по годам
            salary_year_prof(dict{int: int}): Зарпалата выбранной профессии по годам
            count_year_prof(dict{int: int}): Колличество вакансий выбранной пофесиий по годам
            salary_dict_by_city(dict{str: int}): Средняя зарплата по городам,
            count_vacancies_dict_by_city(dict{str: int or str}): Количество вакансий по городам
            prof_name(str): Название професии
        Returns:
             plt(matplotlib.pyplot): Темплэйт созданного графика диаграмм
        """
        x = np.arange(16)

        fig, axs = plt.subplots(nrows=2, ncols=2)

        # plot data in grouped manner of bar type
        axs[0, 0].bar(x - 0.25, list(salary_year.values()), width=0.5)
        axs[0, 0].bar(x + 0.25, list(salary_year_prof.values()), width=0.5)
        axs[0, 0].set_xticks(range(len(salary_year)), list(salary_year.keys()), rotation=90)
        axs[0, 0].legend(['средняя з/п', f'з/п {prof_name}'])
        axs[0, 0].grid(visible=True, axis='y')
        axs[0, 0].set_title('Уровень зарплат по годам')

        axs[0, 1].bar(x - 0.25, list(count_year.values()), width=0.5)
        axs[0, 1].bar(x + 0.25, list(count_year_prof.values()), width=0.5)
        axs[0, 1].set_xticks(range(len(count_year)), list(count_year.keys()), rotation=90)
        axs[0, 1].legend(['Количество вакансий', 'Количество вакансий программист'])
        axs[0, 1].grid(visible=True, axis='y')
        axs[0, 1].set_title('Количество Вакансий по годам')

        plt.rc('ytick', labelsize=SMALL_SIZE)

        list_name = []
        for value in salary_dict_by_city.keys():
            if '-' in value:
                list_name.append(value.replace('-', '-\n'))

            elif ' ' in value:
                list_name.append(value.replace(' ', '\n'))
            else:
                list_name.append(value)

        list_val = [x for x in reversed(list(salary_dict_by_city.values()))]
        list_name = [x for x in reversed(list_name)]

        axs[1, 0].barh(list_name, list_val)
        axs[1, 0].grid(visible=True, axis='x')
        axs[1, 0].set_title('Уровень зарплат по городам')
        plt.rc('ytick', labelsize=SIZE)
        count_vacancies_dict_by_city = Report.remove_percent(count_vacancies_dict_by_city)
        plt.rc('xtick', labelsize=SMALL_SIZE)
        axs[1, 1].pie(list(count_vacancies_dict_by_city.values()), labels=list(count_vacancies_dict_by_city.keys()))
        axs[1, 1].axis('equal')
        axs[1, 1].set_title('Доля вакансий по городам')
        fig.tight_layout()

        return plt

    @staticmethod
    def remove_percent(diction):
        """Функция удаления процентов из строки для построения диаграммы количества вакансй по городам
        Attributes:
            diction(dict{str: int or str}): Необходимый словаь в котором нужно удалить проценты
        Returns:
            dict: Словарь в котором удалены проценты
        """
        return dict((k, float(v[:-1])) for k, v in diction.items())

    @staticmethod
    def generate_pdf(profession, header_1, header_2, rows_1, rows_2, choosing_type):
        """Функция создания PDF файла под нужный метод печати
        Attributes:
            profession(str): Название профессии
            header_1(list[str]): Название колонок первой таблицы
            header_2(list[str]): Название колонок второй таблицы
            rows_1(list[list[str]]): Строки для первой таблицы
            rows_2(list[list[str]]): Строки для второй таблицы
            choosing_type(str): Метод печати выбранный пользователем
        Returns:
            pdf_template(str): Темплэйт для создания PDF файла
        """
        image_file = 'graph.png'
        environment = Environment(loader=FileSystemLoader('.'))
        if choosing_type == 'Вакансии':
            temp = environment.get_template('vacancies_template.html')
            pdf_template = temp.render({
                'title': 'Аналитика по зарплатам и городам для профессии ' + profession,
                'years_title': 'Статистика по годам',
                'years_headers': header_1,
                'years_rows': rows_1,
                'cities_title': 'Статистика по городам',
                'cities_headers': header_2,
                'count_columns': len(header_2),
                'cities_rows': rows_2,
            })
        elif choosing_type == 'Статистика':
            temp = environment.get_template('statistic_template.html')
            pdf_template = temp.render({
                'title': 'Аналитика по зарплатам и городам для профессии ' + profession,
                'image_file': image_file,
            })
        else:
            temp = environment.get_template('default_template.html')
            pdf_template = temp.render({
                'title': 'Аналитика по зарплатам и городам для профессии ' + profession,
                'image_file': image_file,
                'years_title': 'Статистика по годам',
                'years_headers': header_1,
                'years_rows': rows_1,
                'cities_title': 'Статистика по городам',
                'cities_headers': header_2,
                'count_columns': len(header_2),
                'cities_rows': rows_2,
            })
        return pdf_template

class Information:
    """Класс для получения всей нужной информации для создания графиков и таблиц"""
    @staticmethod
    def currency_in_rubles(salary, currency):
        """Функция для перевода зарубежных валют в рубли при помощи словаря currency_to_rub
        Attributes:
            salary(float): Средняя зарплата по вакансиям или выбранной профессии
            currency(str): Валюта вакансии
        Returns:
            float: Средняя зарплата в рублях
        """
        return salary * currency_to_rub[currency]

    @staticmethod
    def get_average_salary(row):
        """Функция подсчёта средней зарплаты в нужной валюте
        Attributes:
            row(list[str]): Вакансия из датасета
        Returns:
            float: Средняя зарплата в рублях
        """
        salary_min, salary_max, salary_currency = float(row.salary.salary_from), float(
            row.salary.salary_to), row.salary.salary_currency
        average_salary = (salary_min + salary_max) / 2
        return Information.currency_in_rubles(average_salary, salary_currency)

    @staticmethod
    def parser_vacancies_by_year(data, years, profession):
        """Функция подсчёта статистики по годам
        Attributes:
            data(list[list[str]]): Вакансии в нужном виде нужных столбцов
            years(int): Год публикации вакансии для сравнения попадает ли вакансия в установленное время
            profession(str): Название професии
        Returns:
             salary_arr(int): Средняя зарплата за год(years)
             length_salary_arr(int): Колличество вакансий за год(years)
             salary_arr_prof(int): Средняя зарплата по выбранной профессии за год(years)
             length_salary_arr_prof(int): Колличество вакансий выбранной профессии за год(years)
        """
        salary_arr = [Information.get_average_salary(row) for row in data if str(years) in row.published_at]
        salary_arr_prof = [Information.get_average_salary(row) for row in data if
                           str(years) in row.published_at and profession in row.name]
        length_salary_arr = len(salary_arr)
        salary_arr = math.floor(sum(salary_arr) / length_salary_arr)
        length_salary_arr_prof = len(salary_arr_prof)
        if length_salary_arr_prof != 0:
            salary_arr_prof = int(sum(salary_arr_prof) / length_salary_arr_prof)
        else:
            salary_arr_prof = 0
        return salary_arr, length_salary_arr, salary_arr_prof, length_salary_arr_prof

    @staticmethod
    def get_city_and_year(data):
        """Функция получения всех лет в датасете, а также получение всех городов в датасете
        Attributes:
            data(list[list[str]]): Вакансии в нужном виде нужных столбцов
        Returns:
            years(dict{int: int}): Словарь, где в качестве ключей перечислены все года из датасета
            cities(dict{str: int}): Словарь, где в качествет ключей перечислены все города из датасета
        """
        cities = {}
        years = {}
        for row in data:
            if row.published_at[:4] not in years.keys():
                years[int(row.published_at[:4])] = 0
            if row.area_name not in cities.keys():
                cities[row.area_name] = 0
            else:
                continue
        return years, cities

    @staticmethod
    def parser_vacancies_by_city(data, cities):
        """Функция подсчёта статистики по городам
        Attributes:
            data(list[list[str]]): Вакансии в нужном виде нужных столбцов
            cities(str): Город для сравнения попадает ли вакансия в нужный регион
        Returns:
            float: Средняя зарплата по городам
            float: Доля вакансий на город
        """
        temp_arr = [Information.get_average_salary(row) for row in data if cities == row.area_name]
        if math.floor(len(temp_arr) / len(data) * 100) >= 1:
            return int((sum(temp_arr)) / len(temp_arr)), len(temp_arr) / len(data)
        else:
            return 0, 0

    @staticmethod
    def work(data, profession):
        """Функция создания статистики для вывода, а также сортировки для удобства чтения
        Attributes:
            data(list[list[str]]): Вакансии в нужном виде нужных столбцов
            profession(str): Название професии
        Returns:
            salary_year(dict{int: int}): Словарь со средними зарплатами по годам
            count_year(dict{int: int}): Словарь с количеством вакансий по годам
            salary_year_prof(dict{int: int}): Словарь со средними зарплатами выбранной профессии по годам
            count_year_prof(dict{int: int}): Словарь с количеством вакансий по выбранной профессии по годам
            salary_dict_by_city(dict{str: int}): Словарь со средними зарплатами по городам
            count_vacancies_dict_by_city(dict{str: int or str}): Словарь с количеством вакансий по городам
        """
        years, cities = Information.get_city_and_year(data)
        salary_year = {}
        count_year = {}
        salary_year_prof = {}
        count_year_prof = {}
        salary_city = {}
        count_city = {}
        for year in years.keys():
            count_vacancies, length_salary_arr, count_vacancies_prof, length_salary_arr_prof = \
                Information.parser_vacancies_by_year(data, year, profession)
            salary_year[year], count_year[year], salary_year_prof[year], count_year_prof[year] = \
                count_vacancies, length_salary_arr, count_vacancies_prof, length_salary_arr_prof
        for city in cities.keys():
            salary_arr_by_city, count_arr_by_city = Information.parser_vacancies_by_city(data, city)
            salary_city[city], count_city[city] = salary_arr_by_city, count_arr_by_city
        salary_dict_by_city = dict(sorted(salary_city.items(), key=lambda x: (-x[1], x[0]))[:10])

        count_vacancies_dict_by_city = dict(sorted(count_city.items(), key=lambda x: -x[1])[:10])
        count_vacancies_dict_by_city = {key: round(count_vacancies_dict_by_city[key], 4) for key in
                                        count_vacancies_dict_by_city}
        salary_dict_by_city = {key: value for key, value in salary_dict_by_city.items() if value != 0}
        count_vacancies_dict_by_city = {key: value for key, value in count_vacancies_dict_by_city.items() if value != 0}
        return salary_year, count_year, salary_year_prof, count_year_prof, salary_dict_by_city, \
               count_vacancies_dict_by_city

    # @staticmethod
    # def go():
    #     file_name = input('Введите название файла: ')
    #     prof_name = input('Введите название профессии: ')
    #     data_set = DataSet(file_name)
    #     salary_year, count_year, salary_year_prof, count_year_prof, salary_dict_by_city, count_vacancies_dict_by_city = \
    #         Information.work(data_set.vacancies, prof_name)
    #     sum_count = sum(list(count_vacancies_dict_by_city.values()))
    #     print(f'Динамика уровня зарплат по годам: {salary_year}')
    #     print(f'Динамика количества вакансий по годам: {count_year}')
    #     print(f'Динамика уровня зарплат по годам для выбранной профессии: {salary_year_prof}')
    #     print(f'Динамика количества вакансий по годам для выбранной профессии: {count_year_prof}')
    #     print(f'Уровень зарплат по городам (в порядке убывания): {salary_dict_by_city}')
    #     print(f'Доля вакансий по городам (в порядке убывания): {count_vacancies_dict_by_city}')
    #
    #     count_vacancies_dict_by_city.update(
    #         (x, f'{round(y * 100, 2)}%') for x, y in count_vacancies_dict_by_city.items())
    #
    #     new_count_vacancies_dict_by_city = {'Другие': f'{100 - (sum_count * 100)}%'}
    #     new_count_vacancies_dict_by_city.update(count_vacancies_dict_by_city)
    #
    #     result_book = Report(salary_year, count_year, salary_year_prof, count_year_prof, salary_dict_by_city,
    #                          count_vacancies_dict_by_city, new_count_vacancies_dict_by_city, prof_name)
    #     result_book.generate_excel.save('report.xlsx')
    #
    #     result_book.generate_image.savefig('graph.png')
    #
    #     pdf_template = result_book.generate_pdf
    #     config = pdfkit.configuration(wkhtmltopdf='C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe')
    #     pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={'enable-local-file-access': True})

