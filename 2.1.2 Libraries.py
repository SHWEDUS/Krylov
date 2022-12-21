import csv
import math

import matplotlib
import matplotlib.pyplot as plt
import numpy as np

import openpyxl
from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter

SMALL_SIZE = 6
SIZE = 8
matplotlib.use('TkAgg')
plt.rcParams.update({'font.size': 8})

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}


class DataSet:

    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies = DataSet.parser_csv(file_name)

    @staticmethod
    def parser_csv(file_name):
        title, data = DataSet.csv_reader(file_name)
        result = DataSet.csv_filer(title, data)
        vacancies = DataSet.get_vacancies(result, title)
        return vacancies

    @staticmethod
    def get_vacancies(result, title):
        vacancies = []
        if result:
            for row in result:
                temp = {}
                for i in range(len(row)):
                    temp[title[i]] = row[i]
                vacancies.append(
                    Vacancies(temp['name'], Salary(temp['salary_from'], temp['salary_to'], temp['salary_currency']),
                              temp['area_name'], temp['published_at']))
            return vacancies
        else:
            return vacancies

    @staticmethod
    def csv_reader(file_name):
        with open(file_name, encoding="utf-8-sig") as file:
            reader = csv.reader(file)
            data_arr = list(reader)
            if not data_arr:
                head_arr = data_arr
                return head_arr, data_arr
            else:
                head_arr = data_arr[0]
                data_arr.pop(0)
                return head_arr, data_arr

    @staticmethod
    def csv_filer(list_naming, reader):
        head_array, data_array = list_naming, reader
        if data_array == [] and head_array == []:
            data_array = []
            return data_array
        for i in range(len(data_array)):
            if '' in data_array[i] or len(data_array[i]) < len(head_array):
                data_array[i].clear()
        data_array = list(filter(None, data_array))

        return data_array


class Salary:
    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency


class Vacancies:
    def __init__(self, name, salary, area_name, published_at):
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at


class Report:
    def __init__(self, salary_year, count_year, salary_year_prof, count_year_prof, salary_dict_by_city,
                 count_vacancies_dict_by_city, new_count, prof_name):
        self.salary_year = salary_year
        self.count_year = count_year
        self.salary_year_prof = salary_year_prof
        self.count_year_prof = count_year_prof
        self.salary_dict_by_city = salary_dict_by_city
        self.count_vacancies_dict_by_city = count_vacancies_dict_by_city
        self.generate_excel = Report.generate_excel(salary_year, count_year, salary_year_prof, count_year_prof,
                                                    salary_dict_by_city, count_vacancies_dict_by_city, prof_name)
        self.generate_image = Report.generate_image(salary_year, count_year, salary_year_prof, count_year_prof,
                                                    salary_dict_by_city, new_count, prof_name)

    @staticmethod
    def generate_excel(salary_year, count_year, salary_year_prof, count_year_prof, salary_dict_by_city,
                       count_vacancies_dict_by_city, profession):
        book = openpyxl.Workbook()
        sheet_year = book.active
        sheet_year.title = 'Статистика по годам'
        sheet_year['A1'] = 'Год'
        sheet_year['B1'] = 'Средняя зарплата'
        sheet_year['C1'] = f'Средняя зарплата - {profession}'
        sheet_year['D1'] = 'Количество вакансий'
        sheet_year['E1'] = f'Количество вакансий - {profession}'

        column_letter_year = 'ABCDE'
        for letter in column_letter_year:
            sheet_year[f'{letter}1'].font = Font(bold=True)

        sheet_city = book.create_sheet('Статистика по городам')
        sheet_city['A1'] = 'Город'
        sheet_city['B1'] = 'Уровень зарплат'
        sheet_city['D1'] = 'Город'
        sheet_city['E1'] = 'Доля вакансий'

        column_letter_city = 'ABDE'
        for letter in column_letter_city:
            sheet_city[f'{letter}1'].font = Font(bold=True)

        for row, (key, value) in enumerate(salary_year.items(), start=2):
            sheet_year[f'A{row}'] = key
        year = [salary_year, salary_year_prof, count_year, count_year_prof]
        Report.fill_column(sheet_year, year, column_letter_year[1:])

        for row, (key, value) in enumerate(salary_dict_by_city.items(), start=2):
            sheet_city[f'A{row}'] = key
            sheet_city[f'D{row}'] = key
            sheet_city[f'B{row}'] = value

        for row, (key, value) in enumerate(count_vacancies_dict_by_city.items(), start=2):
            sheet_city[f'E{row}'] = value

        thin = Side(border_style='thin', color='000000')

        flag = True
        Report.get_width_cells(sheet_year, flag)
        flag = False
        Report.get_width_cells(sheet_city, flag)

        Report.get_border(sheet_year, thin)
        Report.get_border(sheet_city, thin)

        return book

    @staticmethod
    def fill_column(sheet, data, letters):
        for i, item in enumerate(data):
            for row, (key, value) in enumerate(item.items(), start=2):
                sheet[f'{letters[i]}{row}'] = value

    @staticmethod
    def get_width_cells(sheet, flag):
        column_widths = []
        for row in sheet:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell.value)) > column_widths[i]:
                        column_widths[i] = len(str(cell.value))
                else:
                    column_widths += [len(str(cell.value))]

        for i, column_width in enumerate(column_widths, 1):
            if flag:
                sheet.column_dimensions[get_column_letter(i)].width = (column_width + 1)
            else:
                if get_column_letter(i) == 'C':
                    sheet.column_dimensions[get_column_letter(i)].width = 1.5
                else:
                    sheet.column_dimensions[get_column_letter(i)].width = (column_width + 2)

    @staticmethod
    def get_border(sheet, font):
        for row in sheet.columns:
            for cell in row:
                if cell.value:
                    cell.border = Border(top=font, left=font, right=font, bottom=font)

    @staticmethod
    def generate_image(salary_year, count_year, salary_year_prof, count_year_prof,
                       salary_dict_by_city, count_vacancies_dict_by_city, prof_name):

        x = np.arange(16)

        width = 0.2

        fig, axs = plt.subplots(nrows=2, ncols=2)

        # plot data in grouped manner of bar type
        axs[0, 0].bar(x - 0.25, list(salary_year.values()), width=0.5)
        axs[0, 0].bar(x + 0.25, list(salary_year_prof.values()), width=0.5)
        axs[0, 0].set_xticks(range(len(salary_year)), list(salary_year.keys()), rotation=90)
        axs[0, 0].legend(['средняя з/п', f'з/п {prof_name}'])
        axs[0, 0].grid(visible=True, axis='y')
        axs[0, 0].set_title('Уровень зарплат по годам')

        axs[0, 1].bar(x - 0.25, list(count_year.values()), width=0.5)
        axs[0, 1].bar(x + 0.25, list(count_year_prof.values()), width=0.5)
        axs[0, 1].set_xticks(range(len(count_year)), list(count_year.keys()), rotation=90)
        axs[0, 1].legend(['Количество вакансий', 'Количество вакансий программист'])
        axs[0, 1].grid(visible=True, axis='y')
        axs[0, 1].set_title('Количество Вакансий по годам')

        plt.rc('ytick', labelsize=SMALL_SIZE)

        list_name = []
        for value in salary_dict_by_city.keys():
            if '-' in value:
                list_name.append(value.replace('-', '-\n'))

            elif ' ' in value:
                list_name.append(value.replace(' ', '\n'))
            else:
                list_name.append(value)

        list_val = [x for x in reversed(list(salary_dict_by_city.values()))]
        list_name = [x for x in reversed(list_name)]

        axs[1, 0].barh(list_name, list_val)
        axs[1, 0].grid(visible=True, axis='x')
        axs[1, 0].set_title('Уровень зарплат по городам')
        plt.rc('ytick', labelsize=SIZE)

        plt.rc('xtick', labelsize=SMALL_SIZE)
        axs[1, 1].pie(list(count_vacancies_dict_by_city.values()), labels=list(count_vacancies_dict_by_city.keys()))
        axs[1, 1].axis('equal')
        axs[1, 1].set_title('Доля вакансий по городам')
        fig.tight_layout()

        return plt


class Information:

    @staticmethod
    def currency_in_rubles(salary, currency):
        return salary * currency_to_rub[currency]

    @staticmethod
    def get_average_salary(row):
        salary_min, salary_max, salary_currency = float(row.salary.salary_from), float(
            row.salary.salary_to), row.salary.salary_currency
        average_salary = (salary_min + salary_max) / 2
        return Information.currency_in_rubles(average_salary, salary_currency)

    @staticmethod
    def parser_vacancies_by_year(data, years, profession):
        salary_arr = [Information.get_average_salary(row) for row in data if str(years) in row.published_at]
        salary_arr_prof = [Information.get_average_salary(row) for row in data if
                           str(years) in row.published_at and profession in row.name]
        length_salary_arr = len(salary_arr)
        count_vacancies = math.floor(sum(salary_arr) / length_salary_arr)
        length_salary_arr_prof = len(salary_arr_prof)
        if length_salary_arr_prof != 0:
            count_vacancies_prof = int(sum(salary_arr_prof) / length_salary_arr_prof)
        else:
            count_vacancies_prof = 0
        return count_vacancies, length_salary_arr, count_vacancies_prof, length_salary_arr_prof

    @staticmethod
    def get_city_and_year(data):
        cities = {}
        years = {}
        for row in data:
            if row.published_at[:4] not in years.keys():
                years[int(row.published_at[:4])] = 0
            if row.area_name not in cities.keys():
                cities[row.area_name] = 0
            else:
                continue
        return years, cities

    @staticmethod
    def parser_vacancies_by_city(data, cities):
        temp_arr = [Information.get_average_salary(row) for row in data if cities == row.area_name]
        if math.floor(len(temp_arr) / len(data) * 100) >= 1:
            return int((sum(temp_arr)) / len(temp_arr)), len(temp_arr) / len(data)
        else:
            return 0, 0

    @staticmethod
    def work(data, profession):
        years, cities = Information.get_city_and_year(data)
        salary_year = {}
        count_year = {}
        salary_year_prof = {}
        count_year_prof = {}
        salary_city = {}
        count_city = {}
        for year in years.keys():
            count_vacancies, length_salary_arr, count_vacancies_prof, length_salary_arr_prof = \
                Information.parser_vacancies_by_year(data, year, profession)
            salary_year[year], count_year[year], salary_year_prof[year], count_year_prof[year] = \
                count_vacancies, length_salary_arr, count_vacancies_prof, length_salary_arr_prof
        for city in cities.keys():
            salary_arr_by_city, count_arr_by_city = Information.parser_vacancies_by_city(data, city)
            salary_city[city], count_city[city] = salary_arr_by_city, count_arr_by_city
        salary_dict_by_city = dict(sorted(salary_city.items(), key=lambda x: (-x[1], x[0]))[:10])

        count_vacancies_dict_by_city = dict(sorted(count_city.items(), key=lambda x: -x[1])[:10])
        count_vacancies_dict_by_city = {key: round(count_vacancies_dict_by_city[key], 4) for key in
                                        count_vacancies_dict_by_city}
        salary_dict_by_city = {key: value for key, value in salary_dict_by_city.items() if value != 0}
        count_vacancies_dict_by_city = {key: value for key, value in count_vacancies_dict_by_city.items() if value != 0}
        return salary_year, count_year, salary_year_prof, count_year_prof, salary_dict_by_city, \
               count_vacancies_dict_by_city

    @staticmethod
    def go():
        file_name = input('Введите название файла: ')
        prof_name = input('Введите название профессии: ')
        data_set = DataSet(file_name)
        salary_year, count_year, salary_year_prof, count_year_prof, salary_dict_by_city, count_vacancies_dict_by_city = \
            Information.work(data_set.vacancies, prof_name)
        sum_count = sum(list(count_vacancies_dict_by_city.values()))
        print(f'Динамика уровня зарплат по годам: {salary_year}')
        print(f'Динамика количества вакансий по годам: {count_year}')
        print(f'Динамика уровня зарплат по годам для выбранной профессии: {salary_year_prof}')
        print(f'Динамика количества вакансий по годам для выбранной профессии: {count_year_prof}')
        print(f'Уровень зарплат по городам (в порядке убывания): {salary_dict_by_city}')
        print(f'Доля вакансий по городам (в порядке убывания): {count_vacancies_dict_by_city}')

        count_vacancies_dict_by_city.update(
            (x, f'{round(y * 100, 2)}%') for x, y in count_vacancies_dict_by_city.items())

        new_count_vacancies_dict_by_city = count_vacancies_dict_by_city
        new_count_vacancies_dict_by_city['Другие'] = f'{100 - (sum_count * 100)}%'
        new_count_vacancies_dict_by_city.update((x, float(y[:-1])) for x, y in new_count_vacancies_dict_by_city.items()
                                                if '%' in y)

        result_book = Report(salary_year, count_year, salary_year_prof, count_year_prof, salary_dict_by_city,
                             count_vacancies_dict_by_city, new_count_vacancies_dict_by_city, prof_name)
        result_book.generate_excel.save('report.xlsx')

        result_book.generate_image.savefig('graph.png')


Information().go()
